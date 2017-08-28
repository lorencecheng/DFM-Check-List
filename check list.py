# test the logic of github
'''
    20170827: I need to try to code to load all sheet from my original Excel check list, it is seperated by different sheet.
    20170828: I try to move the database excle file to another folder, then let program judge whether the folder exist
'''

import openpyxl
import os

# define a function to print any sheet items
def print_check_point(sheet_name):
    ws=wb.get_sheet_by_name(sheet_name)
    i=ws.max_row
    for row in range(2,i+1):
        print(str(ws["A"+str(row)].value)," : " , str(ws["B"+str(row)].value))

#Start the main code
current_path=os.getcwd()
print("Python Program in this folder: " + str(current_path))
print("")
database_folder="C:/Users/Lorence/Desktop/excel test"  # place the Excel file in this folder, need update if the file location change.
try:
    if os.path.exists(database_folder):
        os.chdir(database_folder)
    wb=openpyxl.load_workbook("database checking items.xlsx",data_only=True)
    sheet_name=[]
    for sheet_name in wb.get_sheet_names():
        print_check_point(sheet_name)
        print("----------------------------------------"+str(sheet_name) + "-----Finish---------------------------------------")
        print("")

except IOError:    #if no this excel folder, alarm
    print("Sorry, I can't find the database Excel file! Please check folder: " + database_folder)
